"""Focused MO-S.6D-SLICE8.1A parser and read-only preview tests."""

import copy
from datetime import date, datetime, time, timedelta
from io import BytesIO
import os
from pathlib import Path
import warnings
from unittest import skipUnless
from unittest.mock import patch
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.admin.models import LogEntry
from django.contrib.auth.models import User
from django.core import signing
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.http import Http404
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from accounts.models import (
    ChurchRoleAssignment,
    ChurchStructureMembership,
    ChurchStructureUnit,
)
from ministry.models import (
    MinistryTeam,
    MinistryTeamParentLink,
    MinistryTeamRoleAssignment,
    MinistryTeamRoleType,
    TeamAssignment,
    TeamAssignmentMember,
    TeamMembership,
)
from ministry.services.worship_xlsx_preview import (
    CONTRACT_REVISION,
    EXPECTED_REAL_WORKBOOK_SHA256,
    EXPECTED_SHEET_NAMES,
    EXPECTED_TITLE,
    INTEGRATION_KEY,
    MAX_UPLOAD_BYTES,
    MAX_ZIP_MEMBERS,
    MAX_ZIP_UNCOMPRESSED_BYTES,
    OBSERVED_REAL_WORKBOOK_TOKEN_COUNTS,
    NORMALIZED_PREVIEW_CONTRACT_REVISION,
    NORMALIZED_PREVIEW_SIGNING_SALT,
    SIGNING_SALT,
    TOKEN_ORDER,
    MappingValidationError,
    PreviewBlocker,
    PreviewClassification,
    SignedWorkbookStateError,
    TargetServiceProfileError,
    TargetServiceProfileErrorCode,
    TargetMatchState,
    WorkbookContractError,
    WorkbookErrorCode,
    build_worship_import_preview,
    decode_parsed_workbook,
    decode_signed_worship_import_preview,
    mapping_candidate_teams,
    match_exact_service_event_targets,
    parse_known_worship_workbook,
    preflight_xlsx_archive,
    sign_parsed_workbook,
)
from notifications.models import Notification

from .models import (
    ServiceEvent,
    ServiceEventAudienceScope,
    ServiceEventPlannerAssignment,
    ServiceEventRequiredTeam,
    ServiceProfile,
)


WORKBOOK_PATH = os.environ.get("SVCA_WORSHIP_WORKBOOK_PATH", "")
_XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DEFAULT_PROFILE = object()


def _supported_rows():
    return [4, *range(6, 57)]


def build_known_workbook(
    *,
    b2=EXPECTED_TITLE,
    a3=None,
    b3="Worship/AV @Bethany",
    missing_sheet=False,
    outside_content=False,
    a4_formula=False,
    friday="1/9/26(Fri)",
    formula_overrides=None,
    cache_overrides=None,
    token_overrides=None,
    tokens=None,
):
    """Create a synthetic strict-contract workbook with real cached formulas."""

    formula_overrides = formula_overrides or {}
    cache_overrides = cache_overrides or {}
    token_overrides = token_overrides or {}
    workbook = Workbook()
    first_title = "Wrong" if missing_sheet else EXPECTED_SHEET_NAMES[0]
    sheet = workbook.active
    sheet.title = first_title
    for title in EXPECTED_SHEET_NAMES[1:]:
        workbook.create_sheet(title)
    sheet["B2"] = b2
    sheet["A3"] = a3
    sheet["B3"] = b3
    sheet.merge_cells("N2:O2")
    sheet["A4"] = "=1+1" if a4_formula else date(2026, 1, 4)
    sheet["A4"].number_format = "yyyy-mm-dd"
    sheet["A5"] = friday

    tokens = tokens or [
        *("A" for _ in range(12)),
        *("C1" for _ in range(13)),
        *("C2" for _ in range(13)),
        *("C3" for _ in range(14)),
    ]
    if len(tokens) != len(_supported_rows()):
        raise ValueError("Synthetic workbook token sequence must contain 52 rows.")
    for index, row_number in enumerate(_supported_rows()):
        if row_number != 4:
            prior = 4 if row_number == 6 else row_number - 1
            sheet.cell(row_number, 1).value = formula_overrides.get(
                row_number, f"=A{prior}+7"
            )
            sheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
        token = token_overrides.get(row_number, tokens[index])
        sheet.cell(row_number, 2).value = f"{token}-Private Leader {row_number}"

    for row_number in (57, 58):
        sheet.cell(row_number, 1).value = formula_overrides.get(
            row_number, f"=A{row_number - 1}+7"
        )
        sheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
    if outside_content:
        sheet["P59"] = "contract drift"

    raw = BytesIO()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook.save(raw)

    source = ZipFile(BytesIO(raw.getvalue()), "r")
    xml = ElementTree.fromstring(source.read("xl/worksheets/sheet1.xml"))
    cells = {
        cell.attrib["r"]: cell
        for cell in xml.findall(f".//{{{_XML_NS}}}c")
        if "r" in cell.attrib
    }
    for row_number in range(6, 59):
        coordinate = f"A{row_number}"
        cell = cells[coordinate]
        cached_value = cache_overrides.get(
            row_number, date(2026, 1, 4) + timedelta(weeks=row_number - 5)
        )
        value_node = cell.find(f"{{{_XML_NS}}}v")
        if value_node is None:
            value_node = ElementTree.SubElement(cell, f"{{{_XML_NS}}}v")
        value_node.text = None if cached_value is None else str(to_excel(cached_value))

    ElementTree.register_namespace("", _XML_NS)
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = ElementTree.tostring(xml, encoding="utf-8", xml_declaration=True)
            target.writestr(info, data)
    source.close()
    return output.getvalue()


def replace_zip_member(content, member_name, replacement):
    source = ZipFile(BytesIO(content), "r")
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            target.writestr(
                info,
                replacement if info.filename == member_name else data,
            )
    source.close()
    return output.getvalue()


def mark_first_zip_member_encrypted(content):
    marked = bytearray(content)
    local_offset = marked.index(b"PK\x03\x04")
    local_flags = int.from_bytes(marked[local_offset + 6 : local_offset + 8], "little")
    marked[local_offset + 6 : local_offset + 8] = (local_flags | 0x1).to_bytes(
        2, "little"
    )
    central_offset = marked.index(b"PK\x01\x02")
    central_flags = int.from_bytes(
        marked[central_offset + 8 : central_offset + 10], "little"
    )
    marked[central_offset + 8 : central_offset + 10] = (
        central_flags | 0x1
    ).to_bytes(2, "little")
    return bytes(marked)


class WorshipWorkbookParserTests(SimpleTestCase):
    def test_valid_known_contract_uses_literal_and_cached_formula_dates(self):
        parsed = parse_known_worship_workbook(
            build_known_workbook(), filename="known.xlsx"
        )
        self.assertEqual(len(parsed.rows), 52)
        self.assertEqual(parsed.rows[0].local_date, date(2026, 1, 4))
        self.assertEqual(parsed.rows[0].date_kind, "literal")
        self.assertEqual(parsed.rows[-1].local_date, date(2026, 12, 27))
        self.assertEqual(parsed.rows[-1].date_kind, "formula_cached")
        self.assertEqual(
            parsed.token_counts, OBSERVED_REAL_WORKBOOK_TOKEN_COUNTS
        )
        self.assertEqual([item["row"] for item in parsed.unsupported_rows], [5, 57, 58])
        self.assertNotIn("Private Leader", repr(parsed))

    @skipUnless(WORKBOOK_PATH and Path(WORKBOOK_PATH).is_file(), "real workbook not supplied")
    def test_real_workbook_acceptance(self):
        path = Path(WORKBOOK_PATH)
        content = path.read_bytes()
        resources = preflight_xlsx_archive(content)
        parsed = parse_known_worship_workbook(content, filename=path.name)
        self.assertEqual(path.stat().st_size, 257609)
        self.assertEqual(resources.member_count, 46)
        self.assertEqual(resources.total_uncompressed_bytes, 2291811)
        self.assertEqual(resources.largest_member_bytes, 631391)
        self.assertEqual(parsed.sha256, EXPECTED_REAL_WORKBOOK_SHA256)
        self.assertEqual(len(parsed.rows), 52)
        self.assertEqual(parsed.rows[0].local_date, date(2026, 1, 4))
        self.assertEqual(parsed.rows[-1].local_date, date(2026, 12, 27))
        self.assertEqual(
            sum(row.date_kind == "formula_cached" for row in parsed.rows), 51
        )
        self.assertEqual(
            parsed.token_counts, OBSERVED_REAL_WORKBOOK_TOKEN_COUNTS
        )

    def test_allowed_token_distribution_is_derived_not_frozen(self):
        tokens = [*("A" for _ in range(12)), *("C1" for _ in range(12))]
        tokens += [*("C2" for _ in range(14)), *("C3" for _ in range(14))]
        parsed = parse_known_worship_workbook(
            build_known_workbook(tokens=tokens)
        )
        self.assertEqual(
            parsed.token_counts,
            {"A": 12, "C1": 12, "C2": 14, "C3": 14},
        )

    def test_allowed_token_may_be_absent(self):
        tokens = [*("A" for _ in range(12)), *("C2" for _ in range(20))]
        tokens += [*("C3" for _ in range(20))]
        parsed = parse_known_worship_workbook(
            build_known_workbook(tokens=tokens)
        )
        self.assertEqual(parsed.token_counts, {"A": 12, "C2": 20, "C3": 20})
        self.assertNotIn("C1", parsed.token_counts)

    def test_missing_sheet_title_headers_merge_and_geometry_fail_closed(self):
        cases = (
            (build_known_workbook(missing_sheet=True), WorkbookErrorCode.SHEET_MISSING),
            (build_known_workbook(b2="wrong"), WorkbookErrorCode.HEADER_MISMATCH),
            (build_known_workbook(a3="Date"), WorkbookErrorCode.HEADER_MISMATCH),
            (build_known_workbook(b3="wrong"), WorkbookErrorCode.HEADER_MISMATCH),
            (
                build_known_workbook(outside_content=True),
                WorkbookErrorCode.CONTRACT_MISMATCH,
            ),
        )
        for content, code in cases:
            with self.subTest(code=code), self.assertRaises(WorkbookContractError) as raised:
                parse_known_worship_workbook(content)
            self.assertEqual(raised.exception.code, code)

    def test_malformed_and_encrypted_workbooks_fail_without_raw_exception(self):
        cases = (
            (b"not-a-zip", WorkbookErrorCode.INVALID_XLSX),
            (
                replace_zip_member(
                    build_known_workbook(), "xl/workbook.xml", b"<broken"
                ),
                WorkbookErrorCode.INVALID_XLSX,
            ),
            (bytes.fromhex("D0CF11E0A1B11AE1") + b"encrypted", WorkbookErrorCode.ENCRYPTED_XLSX),
        )
        for content, code in cases:
            with self.subTest(code=code), self.assertRaises(WorkbookContractError) as raised:
                parse_known_worship_workbook(content)
            self.assertEqual(raised.exception.code, code)

    def test_zip_resource_limits_and_encryption_fail_before_openpyxl(self):
        too_many_members = BytesIO()
        with ZipFile(too_many_members, "w", ZIP_DEFLATED) as archive:
            for index in range(MAX_ZIP_MEMBERS + 1):
                archive.writestr(f"member-{index}.xml", b"")
        too_large = BytesIO()
        with ZipFile(too_large, "w", ZIP_DEFLATED) as archive:
            archive.writestr(
                "xl/worksheets/sheet1.xml",
                b"x" * (MAX_ZIP_UNCOMPRESSED_BYTES + 1),
            )
        cases = (
            (too_many_members.getvalue(), WorkbookErrorCode.RESOURCE_LIMIT),
            (too_large.getvalue(), WorkbookErrorCode.RESOURCE_LIMIT),
            (
                mark_first_zip_member_encrypted(build_known_workbook()),
                WorkbookErrorCode.ENCRYPTED_XLSX,
            ),
        )
        for content, code in cases:
            with (
                self.subTest(code=code),
                patch(
                    "ministry.services.worship_xlsx_preview.load_workbook"
                ) as workbook_loader,
                self.assertRaises(WorkbookContractError) as raised,
            ):
                parse_known_worship_workbook(content)
            self.assertEqual(raised.exception.code, code)
            workbook_loader.assert_not_called()

    def test_literal_formula_cache_and_weekly_sequence_are_strict(self):
        variants = (
            build_known_workbook(a4_formula=True),
            build_known_workbook(cache_overrides={6: None}),
            build_known_workbook(cache_overrides={6: date(2026, 1, 18)}),
            build_known_workbook(cache_overrides={6: date(2027, 1, 11)}),
            build_known_workbook(formula_overrides={6: "=A4+8"}),
            build_known_workbook(formula_overrides={6: "2026-01-11"}),
        )
        for content in variants:
            with self.subTest(), self.assertRaises(WorkbookContractError) as raised:
                parse_known_worship_workbook(content)
            self.assertIn(
                raised.exception.code,
                {
                    WorkbookErrorCode.DATE_MISMATCH,
                    WorkbookErrorCode.FORMULA_CACHE_MISMATCH,
                },
            )

    def test_friday_spillover_and_tokens_are_not_fuzzy_operational_input(self):
        parsed = parse_known_worship_workbook(build_known_workbook())
        self.assertNotIn(date(2026, 1, 9), [row.local_date for row in parsed.rows])
        self.assertFalse(any(row.local_date.year == 2027 for row in parsed.rows))
        for token in ("C", "C10", "prefix C1"):
            with self.subTest(token=token), self.assertRaises(WorkbookContractError) as raised:
                parse_known_worship_workbook(
                    build_known_workbook(token_overrides={4: token})
                )
            self.assertEqual(raised.exception.code, WorkbookErrorCode.UNSUPPORTED_TOKEN)
        redistributed = parse_known_worship_workbook(
            build_known_workbook(token_overrides={4: "C1"})
        )
        self.assertEqual(redistributed.token_counts["C1"], 14)
        self.assertEqual(redistributed.token_counts["A"], 11)


class WorshipWorkbookDomainTestBase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.parsed = parse_known_worship_workbook(build_known_workbook())
        cls.target_profile = ServiceProfile.objects.create(
            key="bethany_0930_cm",
            name="Bethany 09:30",
            event_type=ServiceEvent.EVENT_SUNDAY_SERVICE,
        )
        cls.other_profile = ServiceProfile.objects.create(
            key="other_profile",
            name="Other profile",
            event_type=ServiceEvent.EVENT_SUNDAY_SERVICE,
        )
        cls.root = ChurchStructureUnit.objects.create(
            code="ROOT", name="Whole Church", name_en="Whole Church",
            unit_type=ChurchStructureUnit.UNIT_ROOT,
        )
        cls.cm = ChurchStructureUnit.objects.create(
            code="CM", name="Chinese Ministry", name_en="Chinese Ministry",
            unit_type=ChurchStructureUnit.UNIT_MINISTRY_CONTEXT, parent=cls.root,
        )
        cls.em = ChurchStructureUnit.objects.create(
            code="EM", name="English Ministry", name_en="English Ministry",
            unit_type=ChurchStructureUnit.UNIT_MINISTRY_CONTEXT, parent=cls.root,
        )
        cls.cm_pool = cls._pool("CM Worship Pool", cls.cm)
        cls.em_pool = cls._pool("EM Worship Pool", cls.em)
        cls.c1 = cls._team("Worship C1", cls.cm_pool)
        cls.c2 = cls._team("Worship C2", cls.cm_pool)
        cls.e1 = cls._team("Worship E1", cls.em_pool)
        cls.staff = User.objects.create_user("xlsx_staff", password="pw", is_staff=True)
        cls.other = User.objects.create_user("xlsx_other", password="pw")

    @classmethod
    def _pool(cls, name, anchor):
        pool = MinistryTeam.objects.create(
            name=name, name_en=name, is_assignable=False,
            is_worship_rotation_pool=True,
        )
        MinistryTeamParentLink.objects.create(
            child_team=pool, parent_church_unit=anchor, is_primary=True
        )
        return pool

    @classmethod
    def _team(cls, name, pool):
        team = MinistryTeam.objects.create(name=name, name_en=name)
        MinistryTeamParentLink.objects.create(
            child_team=team, parent_team=pool, is_primary=True
        )
        return team

    def event_for_row(
        self,
        row_index=0,
        *,
        service_profile=_DEFAULT_PROFILE,
        compatibility_key=None,
        event_type=ServiceEvent.EVENT_SUNDAY_SERVICE,
        local_time=time(9, 30),
        status=ServiceEvent.STATUS_PUBLISHED,
        audience=None,
        with_audience=True,
        anchor=None,
    ):
        local_value = datetime.combine(
            self.parsed.rows[row_index].local_date, local_time
        )
        selected_profile = (
            self.target_profile
            if service_profile is _DEFAULT_PROFILE
            else service_profile
        )
        if compatibility_key is None:
            compatibility_key = (
                selected_profile.key if selected_profile is not None else ""
            )
        event = ServiceEvent.objects.create(
            title=f"Sunday {row_index}", title_en=f"Sunday {row_index}",
            service_profile=selected_profile,
            service_profile_key=compatibility_key,
            event_type=event_type,
            start_datetime=timezone.make_aware(
                local_value, timezone.get_current_timezone()
            ),
            status=status, rotation_anchor_team=anchor,
        )
        if with_audience:
            ServiceEventAudienceScope.objects.create(
                service_event=event, unit=audience or self.cm
            )
        return event

    def mapping(self, team=None):
        return {token: team or self.c1 for token in self.parsed.token_counts}

    def first_preview_row(self, mapping=None):
        preview = build_worship_import_preview(
            parsed=self.parsed,
            mapping=mapping or self.mapping(),
            user=self.staff,
        )
        return preview, preview.rows[0]


class WorshipWorkbookTargetMatchingTests(WorshipWorkbookDomainTestBase):
    def test_all_52_exact_fk_targets_match(self):
        for index in range(len(self.parsed.rows)):
            self.event_for_row(index)
        matches = match_exact_service_event_targets(self.parsed)
        self.assertEqual(len(matches), 52)
        self.assertTrue(
            all(
                match.state == TargetMatchState.EXACT_TARGET_MATCHED
                for match in matches
            )
        )

    def test_target_profile_missing_inactive_and_wrong_type_block_before_matching(self):
        cases = (
            (
                "missing",
                lambda: ServiceProfile.objects.filter(
                    pk=self.target_profile.pk
                ).delete(),
                TargetServiceProfileErrorCode.MISSING,
            ),
            (
                "inactive",
                lambda: ServiceProfile.objects.filter(
                    pk=self.target_profile.pk
                ).update(is_active=False),
                TargetServiceProfileErrorCode.INACTIVE,
            ),
            (
                "wrong_type",
                lambda: ServiceProfile.objects.filter(
                    pk=self.target_profile.pk
                ).update(event_type=ServiceEvent.EVENT_SPECIAL_MEETING),
                TargetServiceProfileErrorCode.EVENT_TYPE_MISMATCH,
            ),
        )
        for name, mutate, expected_code in cases:
            with self.subTest(name=name), transaction.atomic():
                mutate()
                with self.assertRaises(TargetServiceProfileError) as raised:
                    match_exact_service_event_targets(self.parsed)
                self.assertEqual(raised.exception.code, expected_code)
                transaction.set_rollback(True)

    def test_exact_target_and_same_date_parallel_event_is_never_selected(self):
        target = self.event_for_row()
        parallel = self.event_for_row(service_profile=self.other_profile)
        match = match_exact_service_event_targets(self.parsed)[0]
        self.assertEqual(match.state, TargetMatchState.EXACT_TARGET_MATCHED)
        self.assertEqual(match.event, target)
        self.assertNotEqual(match.event, parallel)
        self.assertEqual(match.parallel_evidence_count, 1)

    def test_profile_identity_failures_and_other_profile_are_explicit(self):
        cases = (
            (
                "legacy_only",
                lambda: self.event_for_row(
                    service_profile=None,
                    compatibility_key="bethany_0930_cm",
                ),
                TargetMatchState.TARGET_EVENT_PROFILE_FK_MISSING,
            ),
            (
                "profileless",
                lambda: self.event_for_row(
                    service_profile=None,
                    compatibility_key="",
                ),
                TargetMatchState.NO_TARGET,
            ),
            (
                "fk_key_mismatch",
                lambda: ServiceEvent.objects.filter(
                    pk=self.event_for_row().pk
                ).update(service_profile_key="wrong"),
                TargetMatchState.TARGET_EVENT_PROFILE_IDENTITY_DRIFT,
            ),
            (
                "fk_blank_key",
                lambda: ServiceEvent.objects.filter(
                    pk=self.event_for_row().pk
                ).update(service_profile_key=""),
                TargetMatchState.TARGET_EVENT_PROFILE_IDENTITY_DRIFT,
            ),
            (
                "fk_event_type_mismatch",
                lambda: ServiceEvent.objects.filter(
                    pk=self.event_for_row().pk
                ).update(event_type=ServiceEvent.EVENT_SPECIAL_MEETING),
                TargetMatchState.TARGET_EVENT_PROFILE_IDENTITY_DRIFT,
            ),
            (
                "other_profile",
                lambda: self.event_for_row(service_profile=self.other_profile),
                TargetMatchState.TARGET_EVENT_OWNED_BY_OTHER_PROFILE,
            ),
        )
        for name, create_case, expected_state in cases:
            with self.subTest(name=name), transaction.atomic():
                create_case()
                match = match_exact_service_event_targets(self.parsed)[0]
                self.assertEqual(match.state, expected_state)
                self.assertNotEqual(
                    match.state, TargetMatchState.EXACT_TARGET_MATCHED
                )
                transaction.set_rollback(True)

    def test_multiple_canonical_fk_targets_remain_blocked(self):
        self.event_for_row()
        self.event_for_row()
        match = match_exact_service_event_targets(self.parsed)[0]
        self.assertEqual(match.state, TargetMatchState.MULTIPLE_EXACT_TARGETS)

    def test_missing_duplicate_wrong_time_lifecycle_and_audience_conflicts(self):
        self.assertEqual(
            match_exact_service_event_targets(self.parsed)[0].state,
            TargetMatchState.NO_TARGET,
        )
        self.event_for_row(local_time=time(10, 0))
        self.assertEqual(
            match_exact_service_event_targets(self.parsed)[0].state,
            TargetMatchState.NO_TARGET,
        )
        ServiceEvent.objects.all().delete()
        self.event_for_row()
        self.event_for_row()
        self.assertEqual(
            match_exact_service_event_targets(self.parsed)[0].state,
            TargetMatchState.MULTIPLE_EXACT_TARGETS,
        )
        ServiceEvent.objects.all().delete()
        for status in (ServiceEvent.STATUS_DRAFT, ServiceEvent.STATUS_CANCELLED):
            event = self.event_for_row(status=status)
            self.assertEqual(
                match_exact_service_event_targets(self.parsed)[0].state,
                TargetMatchState.LIFECYCLE_CONFLICT,
            )
            event.delete()
        self.event_for_row(with_audience=False)
        self.assertEqual(
            match_exact_service_event_targets(self.parsed)[0].state,
            TargetMatchState.AUDIENCE_INVALID_CONFLICT,
        )

    def test_inactive_and_overlapping_audience_use_canonical_readiness(self):
        inactive = ChurchStructureUnit.objects.create(
            code="OLD", name="Old", unit_type=ChurchStructureUnit.UNIT_CUSTOM,
            parent=self.root, is_active=False,
        )
        event = self.event_for_row(with_audience=False)
        ServiceEventAudienceScope.objects.bulk_create(
            [ServiceEventAudienceScope(service_event=event, unit=inactive)]
        )
        match = match_exact_service_event_targets(self.parsed)[0]
        self.assertEqual(match.state, TargetMatchState.AUDIENCE_INVALID_CONFLICT)
        self.assertIn("inactive_audience_units", match.audience_readiness["invalid_reasons"])
        event.delete()
        event = self.event_for_row(with_audience=False)
        ServiceEventAudienceScope.objects.bulk_create(
            [
                ServiceEventAudienceScope(service_event=event, unit=self.root),
                ServiceEventAudienceScope(service_event=event, unit=self.cm),
            ]
        )
        match = match_exact_service_event_targets(self.parsed)[0]
        self.assertIn("ancestor_descendant_overlap", match.audience_readiness["invalid_reasons"])


class WorshipWorkbookMappingAndGovernanceTests(WorshipWorkbookDomainTestBase):
    def test_unresolved_mapping_is_blocked_and_inactive_nonassignable_rejected(self):
        self.event_for_row()
        preview = build_worship_import_preview(
            parsed=self.parsed, mapping={}, user=self.staff
        )
        self.assertEqual(
            preview.rows[0].classification, PreviewClassification.BLOCKED
        )
        self.assertEqual(preview.rows[0].blocker, PreviewBlocker.MAPPING_UNRESOLVED)
        self.assertFalse(preview.mapping_complete)
        inactive = self._team("Inactive", self.cm_pool)
        inactive.is_active = False
        inactive.save()
        nonassignable = MinistryTeam.objects.create(
            name="Container", name_en="Container", is_assignable=False
        )
        for team in (inactive, nonassignable):
            mapping = self.mapping()
            mapping["A"] = team
            with self.subTest(team=team), self.assertRaises(MappingValidationError):
                build_worship_import_preview(
                    parsed=self.parsed, mapping=mapping, user=self.staff
                )

    def test_target_blocker_precedes_unresolved_mapping(self):
        preview = build_worship_import_preview(
            parsed=self.parsed, mapping={}, user=self.staff
        )
        self.assertEqual(preview.rows[0].blocker, PreviewBlocker.TARGET_MISSING)

    def test_no_candidate_union_still_builds_blocked_preview(self):
        self.event_for_row()
        MinistryTeamParentLink.objects.filter(child_team=self.cm_pool).delete()
        candidates = mapping_candidate_teams(self.parsed)
        self.assertEqual(candidates[self.parsed.rows[0].token], ())
        preview = build_worship_import_preview(
            parsed=self.parsed, mapping={}, user=self.staff
        )
        self.assertEqual(preview.rows[0].blocker, PreviewBlocker.MAPPING_UNRESOLVED)

    def test_candidate_controls_use_union_but_each_destination_fails_closed(self):
        self.event_for_row(0, audience=self.cm)
        self.event_for_row(1, audience=self.em)
        candidates = mapping_candidate_teams(self.parsed)
        self.assertIn(self.c1, candidates["A"])
        self.assertIn(self.e1, candidates["A"])
        preview = build_worship_import_preview(
            parsed=self.parsed, mapping=self.mapping(self.c1), user=self.staff
        )
        self.assertEqual(preview.rows[0].classification, PreviewClassification.PROPOSED_CHANGE)
        self.assertEqual(preview.rows[1].classification, PreviewClassification.BLOCKED)
        self.assertEqual(preview.rows[1].blocker, PreviewBlocker.TEAM_INELIGIBLE)

    def test_no_op_initial_change_clean_change_and_current_assignment_block(self):
        scenarios = (
            (self.c1, None, PreviewClassification.NO_OP, None),
            (None, None, PreviewClassification.PROPOSED_CHANGE, None),
            (self.c2, None, PreviewClassification.PROPOSED_CHANGE, None),
            (
                self.c2,
                self.c2,
                PreviewClassification.BLOCKED,
                PreviewBlocker.CURRENT_WORSHIP_ASSIGNMENT,
            ),
        )
        for anchor, assignment_team, classification, blocker in scenarios:
            with self.subTest(anchor=anchor, assignment_team=assignment_team):
                ServiceEvent.objects.all().delete()
                event = self.event_for_row(anchor=anchor)
                if assignment_team:
                    TeamAssignment.objects.bulk_create(
                        [TeamAssignment(service_event=event, ministry_team=assignment_team)]
                    )
                _, row = self.first_preview_row()
                self.assertEqual(row.classification, classification)
                self.assertEqual(row.blocker, blocker)

    def test_off_team_out_of_scope_and_duplicate_ownership_block(self):
        cases = ("off_team", "out_of_scope", "duplicate")
        for case in cases:
            with self.subTest(case=case):
                ServiceEvent.objects.all().delete()
                event = self.event_for_row(anchor=self.c1)
                if case == "off_team":
                    assignments = [TeamAssignment(service_event=event, ministry_team=self.c2)]
                elif case == "out_of_scope":
                    assignments = [TeamAssignment(service_event=event, ministry_team=self.e1)]
                else:
                    assignments = [
                        TeamAssignment(service_event=event, ministry_team=self.c1),
                        TeamAssignment(service_event=event, ministry_team=self.c1),
                    ]
                TeamAssignment.objects.bulk_create(assignments)
                _, row = self.first_preview_row()
                self.assertEqual(row.classification, PreviewClassification.BLOCKED)
                self.assertEqual(row.blocker, PreviewBlocker.OWNERSHIP_CONFLICT)

    def test_downstream_projection_is_roster_free_and_deterministic(self):
        event = self.event_for_row(anchor=None)
        downstream = MinistryTeam.objects.create(
            name="Projection", name_en="Projection"
        )
        ServiceEventRequiredTeam.objects.create(
            service_event=event, ministry_team=downstream
        )
        TeamAssignment.objects.bulk_create(
            [
                TeamAssignment(
                    service_event=event, ministry_team=downstream,
                    status=TeamAssignment.STATUS_PREPARED, notes="Private note",
                )
            ]
        )
        preview, row = self.first_preview_row()
        self.assertEqual(row.downstream_impacts[0].team, downstream)
        self.assertEqual(row.downstream_impacts[0].statuses, ("prepared",))
        signed_text = str(
            decode_signed_worship_import_preview(
                preview.signed_payload, user=self.staff
            )
        )
        self.assertNotIn("Private note", signed_text)
        self.assertNotIn("Private Leader", signed_text)
        self.assertEqual(
            [item.source.source_row for item in preview.rows], _supported_rows()
        )

    def test_signed_state_is_user_bound_expiring_tamper_rejected_and_bounded(self):
        self.event_for_row()
        parsed_token = sign_parsed_workbook(self.parsed, user=self.staff)
        parsed_payload = signing.loads(parsed_token, salt=SIGNING_SALT)
        self.assertEqual(parsed_payload["integration_key"], INTEGRATION_KEY)
        self.assertEqual(
            parsed_payload["target_profile"],
            {
                "profile_id": self.target_profile.pk,
                "profile_key": self.target_profile.key,
                "profile_event_type": self.target_profile.event_type,
            },
        )
        decoded = decode_parsed_workbook(parsed_token, user=self.staff)
        self.assertEqual(decoded.rows, self.parsed.rows)
        with self.assertRaises(SignedWorkbookStateError):
            decode_parsed_workbook(parsed_token, user=self.other)
        with self.assertRaises(SignedWorkbookStateError):
            decode_parsed_workbook(parsed_token + "tampered", user=self.staff)
        with self.assertRaises(SignedWorkbookStateError):
            decode_parsed_workbook(parsed_token, user=self.staff, max_age=-1)
        preview, _ = self.first_preview_row()
        self.assertLess(preview.signed_payload_bytes, 64 * 1024)
        normalized = decode_signed_worship_import_preview(
            preview.signed_payload, user=self.staff
        )
        self.assertEqual(
            normalized["contract_revision"],
            NORMALIZED_PREVIEW_CONTRACT_REVISION,
        )
        self.assertEqual(normalized["integration_key"], INTEGRATION_KEY)
        self.assertEqual(normalized["target_profile"], parsed_payload["target_profile"])
        matched_row = normalized["rows"][0]
        self.assertEqual(matched_row["service_profile_id"], self.target_profile.pk)
        self.assertEqual(matched_row["service_profile_key"], self.target_profile.key)
        self.assertEqual(
            matched_row["service_profile_event_type"],
            self.target_profile.event_type,
        )
        self.assertEqual(matched_row["profile_identity_state"], "exact")

    def test_v1_parsed_and_normalized_tokens_are_rejected(self):
        self.event_for_row()
        parsed_token = sign_parsed_workbook(self.parsed, user=self.staff)
        parsed_payload = signing.loads(parsed_token, salt=SIGNING_SALT)
        parsed_payload["contract_revision"] = "SVCA_BETHANY_0930_2026_V1"
        parsed_payload["signing_version"] = 1
        v1_parsed = signing.dumps(
            parsed_payload,
            compress=True,
            salt="ministry.worship-xlsx-preview.v1",
        )
        with self.assertRaises(SignedWorkbookStateError):
            decode_parsed_workbook(v1_parsed, user=self.staff)

        preview, _ = self.first_preview_row()
        normalized_payload = signing.loads(
            preview.signed_payload,
            salt=NORMALIZED_PREVIEW_SIGNING_SALT,
        )
        normalized_payload["contract_revision"] = (
            "SVCA_BETHANY_0930_2026_V1"
        )
        normalized_payload["signing_version"] = 1
        v1_normalized = signing.dumps(
            normalized_payload,
            compress=True,
            salt="ministry.worship-xlsx-preview.v1",
        )
        with self.assertRaises(SignedWorkbookStateError):
            decode_signed_worship_import_preview(v1_normalized, user=self.staff)

    def test_v2_signed_shapes_reject_missing_wrong_or_extra_identity(self):
        self.event_for_row()
        parsed_token = sign_parsed_workbook(self.parsed, user=self.staff)
        parsed_base = signing.loads(parsed_token, salt=SIGNING_SALT)
        parsed_cases = []
        for mutation in (
            lambda payload: payload.pop("target_profile"),
            lambda payload: payload.update(integration_key="other_adapter"),
            lambda payload: payload["target_profile"].update(profile_id=999999),
            lambda payload: payload["target_profile"].update(profile_key="wrong"),
            lambda payload: payload["target_profile"].update(
                profile_event_type=ServiceEvent.EVENT_SPECIAL_MEETING
            ),
            lambda payload: payload.update(unexpected=True),
        ):
            payload = copy.deepcopy(parsed_base)
            mutation(payload)
            parsed_cases.append(
                signing.dumps(payload, compress=True, salt=SIGNING_SALT)
            )
        for token in parsed_cases:
            with self.subTest(token=token[:12]), self.assertRaises(
                SignedWorkbookStateError
            ):
                decode_parsed_workbook(token, user=self.staff)

        preview, _ = self.first_preview_row()
        normalized_base = signing.loads(
            preview.signed_payload,
            salt=NORMALIZED_PREVIEW_SIGNING_SALT,
        )
        normalized_cases = []
        for mutation in (
            lambda payload: payload.pop("target_profile"),
            lambda payload: payload.update(integration_key="other_adapter"),
            lambda payload: payload["target_profile"].update(profile_id=999999),
            lambda payload: payload["target_profile"].update(profile_key="wrong"),
            lambda payload: payload["target_profile"].update(
                profile_event_type=ServiceEvent.EVENT_SPECIAL_MEETING
            ),
            lambda payload: payload["rows"][0].update(service_profile_id=999999),
            lambda payload: payload.update(unexpected=True),
        ):
            payload = copy.deepcopy(normalized_base)
            mutation(payload)
            normalized_cases.append(
                signing.dumps(
                    payload,
                    compress=True,
                    salt=NORMALIZED_PREVIEW_SIGNING_SALT,
                )
            )
        for token in normalized_cases:
            with self.subTest(token=token[:12]), self.assertRaises(
                SignedWorkbookStateError
            ):
                decode_signed_worship_import_preview(token, user=self.staff)

    def test_signed_parsed_state_preserves_counts_and_date_kind_semantics(self):
        tokens = [*("A" for _ in range(12)), *("C1" for _ in range(12))]
        tokens += [*("C2" for _ in range(14)), *("C3" for _ in range(14))]
        parsed = parse_known_worship_workbook(
            build_known_workbook(tokens=tokens)
        )
        signed = sign_parsed_workbook(parsed, user=self.staff)
        decoded = decode_parsed_workbook(signed, user=self.staff)
        self.assertEqual(
            decoded.token_counts,
            {"A": 12, "C1": 12, "C2": 14, "C3": 14},
        )

        cases = (("token_counts", None), ("date_kind", 0), ("date_kind", 1))
        for field, row_index in cases:
            with self.subTest(field=field, row_index=row_index):
                payload = signing.loads(signed, salt=SIGNING_SALT)
                if field == "token_counts":
                    payload["token_counts"]["A"] += 1
                elif row_index == 0:
                    payload["rows"][row_index]["date_kind"] = "formula_cached"
                else:
                    payload["rows"][row_index]["date_kind"] = "literal"
                resigned = signing.dumps(
                    payload, compress=True, salt=SIGNING_SALT
                )
                with self.assertRaises(SignedWorkbookStateError):
                    decode_parsed_workbook(resigned, user=self.staff)


@override_settings(
    CMS_ENABLED_INTEGRATIONS=["svca_bethany_2026_worship_xlsx"]
)
class WorshipWorkbookViewTests(WorshipWorkbookDomainTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.c3 = cls._team("Worship C3", cls.cm_pool)
        cls.team_a = cls._team("Worship A", cls.cm_pool)
        cls.superuser = User.objects.create_user(
            "xlsx_super", password="pw", is_superuser=True, is_staff=False
        )
        cls.ordinary = User.objects.create_user("xlsx_ordinary", password="pw")
        cls.exact_lead = User.objects.create_user("xlsx_exact_lead", password="pw")
        cls.pool_lead = User.objects.create_user("xlsx_pool_lead", password="pw")
        cls.event_planner = User.objects.create_user("xlsx_planner", password="pw")
        cls.global_manager = User.objects.create_user("xlsx_global", password="pw")
        cls.role_type = MinistryTeamRoleType.objects.create(
            code=MinistryTeamRoleType.CODE_LEAD,
            name="Lead", name_en="Lead", is_active=True,
        )
        for user, team in ((cls.exact_lead, cls.c1), (cls.pool_lead, cls.cm_pool)):
            MinistryTeamRoleAssignment.objects.create(
                team=team, role_type=cls.role_type, user=user,
                start_date=date(2025, 1, 1), is_active=True,
            )
        ChurchRoleAssignment.objects.create(
            user=cls.global_manager,
            role=ChurchRoleAssignment.ROLE_COWORKER,
            scope_type=ChurchRoleAssignment.SCOPE_GLOBAL,
        )
        cls.events = []
        for index, row in enumerate(cls.parsed.rows):
            local_value = datetime.combine(row.local_date, time(9, 30))
            event = ServiceEvent.objects.create(
                title=f"Sunday {index}", title_en=f"Sunday {index}",
                service_profile=cls.target_profile,
                service_profile_key="bethany_0930_cm",
                event_type=ServiceEvent.EVENT_SUNDAY_SERVICE,
                start_datetime=timezone.make_aware(
                    local_value, timezone.get_current_timezone()
                ),
                status=(
                    ServiceEvent.STATUS_COMPLETED
                    if row.local_date < date(2026, 8, 27)
                    else ServiceEvent.STATUS_PUBLISHED
                ),
            )
            ServiceEventAudienceScope.objects.create(service_event=event, unit=cls.cm)
            cls.events.append(event)
        ServiceEventPlannerAssignment.objects.create(
            service_event=cls.events[0], user=cls.event_planner
        )

    def upload(self, *, name="schedule.xlsx", content=None):
        return SimpleUploadedFile(
            name,
            content if content is not None else build_known_workbook(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def setUp(self):
        session = self.client.session
        session["language"] = "en"
        session.save()

    def mapping_post(self, token, team=None):
        team = team or self.c1
        data = {"signed_workbook": token}
        for workbook_token in TOKEN_ORDER:
            data[f"mapping_{workbook_token.lower()}"] = team.pk
        return data

    def test_staff_and_superuser_allowed_all_other_authority_classes_denied(self):
        for user in (self.staff, self.superuser):
            with self.subTest(user=user.username):
                self.client.force_login(user)
                self.assertEqual(
                    self.client.get(reverse("worship_workbook_preview")).status_code,
                    200,
                )
        for user in (
            self.ordinary,
            self.exact_lead,
            self.pool_lead,
            self.event_planner,
            self.global_manager,
        ):
            with self.subTest(user=user.username):
                self.client.force_login(user)
                self.assertEqual(
                    self.client.get(reverse("worship_workbook_preview")).status_code,
                    403,
                )

    def test_enabled_staff_sees_workbook_card(self):
        self.client.force_login(self.staff)
        planning = self.client.get(reverse("worship_planning"))
        self.assertEqual(planning.status_code, 200)
        self.assertContains(planning, "Annual Workbook Preview")
        self.assertContains(planning, reverse("worship_workbook_preview"))

    @override_settings(CMS_ENABLED_INTEGRATIONS=[])
    def test_disabled_integration_hides_card_and_staff_or_superuser_cannot_bypass(self):
        for user in (self.staff, self.superuser):
            with self.subTest(user=user.username):
                self.client.force_login(user)
                planning = self.client.get(reverse("worship_planning"))
                self.assertEqual(planning.status_code, 200)
                self.assertContains(planning, reverse("worship_rotation_planner"))
                self.assertNotContains(planning, "Annual Workbook Preview")
                self.assertNotContains(planning, reverse("worship_workbook_preview"))
                self.assertEqual(
                    self.client.get(reverse("worship_workbook_preview")).status_code,
                    404,
                )

    @override_settings(CMS_ENABLED_INTEGRATIONS=[])
    def test_disabled_preview_stops_before_upload_read_parser_service_or_query(self):
        from events.views import worship_workbook_preview

        uploaded = type(
            "UnreadUpload",
            (),
            {
                "name": "schedule.xlsx",
                "size": 10,
                "read": lambda self: (_ for _ in ()).throw(
                    AssertionError("disabled upload was read")
                ),
            },
        )()
        request = RequestFactory().post("/disabled-workbook-preview/")
        request.user = self.staff
        request._files = {"workbook": uploaded}
        with (
            patch(
                "ministry.services.worship_xlsx_preview.parse_known_worship_workbook"
            ) as parser,
            patch(
                "ministry.services.worship_xlsx_preview.ServiceEvent.objects.filter"
            ) as event_query,
            patch("events.forms.WorshipWorkbookUploadForm") as upload_form,
            self.assertNumQueries(0),
            self.assertRaises(Http404),
        ):
            worship_workbook_preview(request)
        parser.assert_not_called()
        event_query.assert_not_called()
        upload_form.assert_not_called()

    def test_wrong_extension_too_large_and_malformed_upload_are_rejected(self):
        self.client.force_login(self.staff)
        cases = (
            (self.upload(name="schedule.xls"), "Only .xlsx"),
            (self.upload(content=b"x" * (MAX_UPLOAD_BYTES + 1)), "5 MiB"),
            (self.upload(content=b"not xlsx"), "Invalid XLSX"),
        )
        for uploaded, text in cases:
            with self.subTest(text=text):
                response = self.client.post(
                    reverse("worship_workbook_preview"), {"workbook": uploaded}
                )
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, text)
                self.assertNotContains(response, "Traceback")

    def test_upload_mapping_preview_is_bilingual_private_and_mints_distinct_confirmation(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            reverse("worship_workbook_preview"), {"workbook": self.upload()}
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, CONTRACT_REVISION)
        self.assertContains(response, "Workbook token A")
        self.assertNotContains(response, "Private Leader")
        token = response.context["mapping_form"]["signed_workbook"].value()
        response = self.client.post(
            reverse("worship_workbook_preview"), self.mapping_post(token)
        )
        self.assertContains(response, "Exact matched targets")
        self.assertContains(response, "52")
        self.assertContains(response, "Complete")
        self.assertContains(response, "Preview only")
        self.assertContains(response, "Apply reviewed Worship Team changes")
        self.assertContains(response, "Confirm and apply 52 reviewed targets")
        self.assertIsNotNone(response.context["confirmation_form"])
        self.assertIsNotNone(response.context["confirmation_proposal"])
        session = self.client.session
        session["language"] = "zh"
        session.save()
        response = self.client.get(reverse("worship_workbook_preview"))
        self.assertContains(response, "仅预览——尚未修改任何排班数据。")
        self.assertContains(response, "上传 XLSX 工作簿")

    def test_preview_renders_scoped_table_usability_hooks_and_explicit_confirm(self):
        self.client.force_login(self.staff)
        upload_response = self.client.post(
            reverse("worship_workbook_preview"), {"workbook": self.upload()}
        )
        token = upload_response.context["mapping_form"]["signed_workbook"].value()
        response = self.client.post(
            reverse("worship_workbook_preview"), self.mapping_post(token)
        )

        self.assertContains(response, 'class="workbook-preview-page"')
        self.assertContains(response, "data-workbook-table-top-scroll")
        self.assertContains(response, "data-workbook-table-top-scroll-spacer")
        self.assertContains(response, "data-workbook-table-scroll")
        self.assertContains(response, "new ResizeObserver")
        self.assertContains(response, "window.requestAnimationFrame")
        self.assertContains(response, "All or nothing")
        self.assertContains(response, "Confirm and apply 52 reviewed targets")

    def test_absent_token_has_no_mapping_control(self):
        self.client.force_login(self.staff)
        tokens = [*("A" for _ in range(12)), *("C2" for _ in range(20))]
        tokens += [*("C3" for _ in range(20))]
        response = self.client.post(
            reverse("worship_workbook_preview"),
            {"workbook": self.upload(content=build_known_workbook(tokens=tokens))},
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("mapping_c1", response.context["mapping_form"].fields)
        self.assertNotContains(response, "Workbook token C1")

    def test_incomplete_mapping_renders_truthful_blocked_preview(self):
        self.client.force_login(self.staff)
        upload_response = self.client.post(
            reverse("worship_workbook_preview"), {"workbook": self.upload()}
        )
        token = upload_response.context["mapping_form"]["signed_workbook"].value()
        data = self.mapping_post(token)
        del data["mapping_c3"]
        response = self.client.post(reverse("worship_workbook_preview"), data)
        preview = response.context["preview"]
        self.assertIsNotNone(preview)
        self.assertEqual(preview.blocked_count, 14)
        self.assertFalse(preview.mapping_complete)
        self.assertContains(response, "Incomplete / blocked")
        self.assertContains(
            response,
            "Mapping unresolved — select an eligible Worship Team.",
        )

    def test_no_eligible_candidates_still_renders_blocked_preview(self):
        self.client.force_login(self.staff)
        MinistryTeamParentLink.objects.filter(child_team=self.cm_pool).delete()
        upload_response = self.client.post(
            reverse("worship_workbook_preview"), {"workbook": self.upload()}
        )
        token = upload_response.context["mapping_form"]["signed_workbook"].value()
        response = self.client.post(
            reverse("worship_workbook_preview"), {"signed_workbook": token}
        )
        self.assertIsNotNone(response.context["preview"])
        self.assertEqual(response.context["preview"].blocked_count, 52)
        self.assertContains(response, "Incomplete / blocked")

    def test_missing_targets_precede_blank_mapping_in_rendered_preview(self):
        self.client.force_login(self.staff)
        missing_dates = {
            row.local_date for row in self.parsed.rows if row.token == "C3"
        }
        for event in self.events:
            if timezone.localtime(event.start_datetime).date() in missing_dates:
                event.delete()
        upload_response = self.client.post(
            reverse("worship_workbook_preview"), {"workbook": self.upload()}
        )
        token = upload_response.context["mapping_form"]["signed_workbook"].value()
        data = self.mapping_post(token)
        del data["mapping_c3"]
        response = self.client.post(reverse("worship_workbook_preview"), data)
        c3_rows = [
            row for row in response.context["preview"].rows if row.source.token == "C3"
        ]
        self.assertTrue(c3_rows)
        self.assertTrue(
            all(row.blocker == PreviewBlocker.TARGET_MISSING for row in c3_rows)
        )
        self.assertContains(response, "No exact target event.")

    def _business_snapshot(self):
        models = (
            ServiceEvent,
            ServiceEventAudienceScope,
            ServiceEventRequiredTeam,
            ServiceEventPlannerAssignment,
            MinistryTeam,
            MinistryTeamParentLink,
            TeamMembership,
            MinistryTeamRoleAssignment,
            TeamAssignment,
            TeamAssignmentMember,
            ChurchStructureUnit,
            ChurchStructureMembership,
            LogEntry,
            Notification,
        )
        return {
            model._meta.label: model.objects.count() for model in models
        } | {
            "event_state": list(
                ServiceEvent.objects.order_by("id").values_list(
                    "id", "scheduling_revision", "service_profile_key",
                    "rotation_anchor_team_id",
                )
            )
        }

    def test_get_valid_preview_and_invalid_preview_write_zero_business_rows(self):
        self.client.force_login(self.staff)
        before = self._business_snapshot()
        with patch("django.db.transaction.on_commit") as on_commit:
            self.client.get(reverse("worship_workbook_preview"))
            upload_response = self.client.post(
                reverse("worship_workbook_preview"), {"workbook": self.upload()}
            )
            token = upload_response.context["mapping_form"]["signed_workbook"].value()
            preview_response = self.client.post(
                reverse("worship_workbook_preview"), self.mapping_post(token)
            )
            invalid_response = self.client.post(
                reverse("worship_workbook_preview"),
                {"workbook": self.upload(content=b"bad")},
            )
        self.assertEqual(preview_response.context["preview"].proposed_change_count, 52)
        self.assertContains(invalid_response, "Invalid XLSX")
        self.assertEqual(self._business_snapshot(), before)
        on_commit.assert_not_called()
