"""Strict SVCA Bethany 09:30 XLSX parsing and read-only preview.

MO-S.6D-SLICE8.1A deliberately supports one code-owned workbook contract.
It evaluates no formulas, persists no upload or proposal, and exposes no
confirmation or scheduling-write operation.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from enum import StrEnum
from hashlib import sha256
from io import BytesIO
import re
from zipfile import BadZipFile, LargeZipFile, ZipFile

from django.core import signing
from django.utils import timezone
from openpyxl import load_workbook

from events.models import ServiceEvent
from events.service_profile_readiness import service_event_audience_readiness

from ..models import MinistryTeam
from .worship_governance import (
    WorshipOwnershipConsistencyState,
    inspect_worship_ownership_consistency,
)
from .worship_rotation_planner import (
    DownstreamImpact,
    build_worship_downstream_projection,
    build_worship_preview_fingerprints,
)


CONTRACT_REVISION = "SVCA_BETHANY_0930_2026_V1"
EXPECTED_REAL_WORKBOOK_SHA256 = (
    "186735DC723979AA49D209C92D4155BE533D6AFE9253CDB5D8B809A77C8B07AA"
)
SUPPORTED_SHEET = "All 930"
EXPECTED_SHEET_NAMES = (
    "All 930",
    "2026 Special Events ",
    "Bethany Hall 其他服事（總務，關懷，福音）",
    "工作表22",
    "三谷主日（擘餅、worship, 总务，饭食，接待，主日學）",
    "Kephir",
    " Sunday Groups ",
    "Children (Auto Renew) ",
)
EXPECTED_TITLE = "2026 SVCA Sunday Service and Special Events Schedule (Master 總表）"
EXPECTED_DATE_HEADER = None
EXPECTED_WORSHIP_HEADER = "Worship/AV @Bethany"
EXPECTED_MERGES = ("N2:O2",)
SUPPORTED_YEAR = 2026
SUPPORTED_PROFILE_KEY = "bethany_0930_cm"
SUPPORTED_LOCAL_TIME = time(9, 30)
SUPPORTED_EVENT_TYPE = ServiceEvent.EVENT_SUNDAY_SERVICE
SUPPORTED_ROWS = (4, *range(6, 57))
TOKEN_ORDER = ("A", "C1", "C2", "C3")
OBSERVED_REAL_WORKBOOK_TOKEN_COUNTS = {
    "A": 12,
    "C1": 13,
    "C2": 13,
    "C3": 14,
}
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_ZIP_MEMBERS = 128
MAX_ZIP_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
MAX_ZIP_MEMBER_UNCOMPRESSED_BYTES = 8 * 1024 * 1024
SIGNING_VERSION = 1
SIGNING_SALT = "ministry.worship-xlsx-preview.v1"
SIGNING_MAX_AGE_SECONDS = 1800

_TOKEN_RE = re.compile(r"^(A|C1|C2|C3)(?=$|[\s\-(])")
_ENCRYPTED_OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")


class WorkbookErrorCode(StrEnum):
    INVALID_XLSX = "invalid_xlsx"
    ENCRYPTED_XLSX = "encrypted_xlsx"
    CONTRACT_MISMATCH = "contract_mismatch"
    SHEET_MISSING = "sheet_missing"
    HEADER_MISMATCH = "header_mismatch"
    DATE_MISMATCH = "date_mismatch"
    FORMULA_CACHE_MISMATCH = "formula_cache_mismatch"
    UNSUPPORTED_TOKEN = "unsupported_token"
    RESOURCE_LIMIT = "resource_limit"


class WorkbookContractError(ValueError):
    def __init__(self, code: WorkbookErrorCode, detail: str):
        self.code = code
        self.detail = detail
        super().__init__(detail)


class SignedWorkbookStateError(ValueError):
    pass


class MappingValidationError(ValueError):
    pass


class TargetMatchState(StrEnum):
    EXACT_TARGET_MATCHED = "exact_target_matched"
    NO_TARGET = "no_target"
    MULTIPLE_EXACT_TARGETS = "multiple_exact_targets"
    LIFECYCLE_CONFLICT = "lifecycle_conflict"
    AUDIENCE_INVALID_CONFLICT = "audience_invalid_conflict"


class PreviewClassification(StrEnum):
    NO_OP = "no_op"
    PROPOSED_CHANGE = "proposed_change"
    BLOCKED = "blocked"


class PreviewBlocker(StrEnum):
    TARGET_MISSING = "target_missing"
    TARGET_AMBIGUOUS = "target_ambiguous"
    TARGET_LIFECYCLE = "target_lifecycle"
    TARGET_AUDIENCE = "target_audience"
    MAPPING_UNRESOLVED = "mapping_unresolved"
    TEAM_INELIGIBLE = "team_ineligible"
    OWNERSHIP_CONFLICT = "ownership_conflict"
    CURRENT_WORSHIP_ASSIGNMENT = "current_worship_assignment"


@dataclass(frozen=True)
class ParsedWorkbookRow:
    source_row: int
    source_cell: str
    local_date: date
    date_kind: str
    token: str


@dataclass(frozen=True)
class ParsedWorshipWorkbook:
    filename: str
    sha256: str
    rows: tuple[ParsedWorkbookRow, ...]
    token_counts: dict[str, int]
    unsupported_rows: tuple[dict, ...]


@dataclass(frozen=True)
class XlsxArchiveResources:
    member_count: int
    total_uncompressed_bytes: int
    largest_member_bytes: int


@dataclass(frozen=True)
class TargetMatch:
    row: ParsedWorkbookRow
    state: TargetMatchState
    event: ServiceEvent | None
    exact_target_ids: tuple[int, ...]
    audience_readiness: dict | None
    parallel_evidence_count: int


@dataclass(frozen=True)
class WorshipImportPreviewRow:
    source: ParsedWorkbookRow
    target_state: TargetMatchState
    event: ServiceEvent | None
    current_team: MinistryTeam | None
    proposed_team: MinistryTeam | None
    classification: PreviewClassification
    blocker: PreviewBlocker | None
    ownership_state: WorshipOwnershipConsistencyState | None
    downstream_impacts: tuple[DownstreamImpact, ...]
    parallel_evidence_count: int
    fingerprints: dict | None


@dataclass(frozen=True)
class WorshipImportPreview:
    parsed: ParsedWorshipWorkbook
    mappings: dict[str, MinistryTeam]
    rows: tuple[WorshipImportPreviewRow, ...]
    normalized_payload: dict
    signed_payload: str

    @property
    def no_op_count(self):
        return sum(
            row.classification == PreviewClassification.NO_OP for row in self.rows
        )

    @property
    def proposed_change_count(self):
        return sum(
            row.classification == PreviewClassification.PROPOSED_CHANGE
            for row in self.rows
        )

    @property
    def blocked_count(self):
        return sum(
            row.classification == PreviewClassification.BLOCKED
            for row in self.rows
        )

    @property
    def matched_target_count(self):
        return sum(
            row.target_state == TargetMatchState.EXACT_TARGET_MATCHED
            for row in self.rows
        )

    @property
    def signed_payload_bytes(self):
        return len(self.signed_payload.encode("utf-8"))

    @property
    def mapping_complete(self):
        if any(self.mappings.get(token) is None for token in self.parsed.token_counts):
            return False
        return not any(
            row.blocker
            in {PreviewBlocker.MAPPING_UNRESOLVED, PreviewBlocker.TEAM_INELIGIBLE}
            for row in self.rows
        )


def _safe_filename(filename):
    return str(filename or "workbook.xlsx").replace("\\", "/").rsplit("/", 1)[-1]


def _expected_date_for_row(row_number):
    index = 0 if row_number == 4 else row_number - 5
    return date(2026, 1, 4) + timedelta(weeks=index)


def _normalize_excel_date(value, *, cell):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise WorkbookContractError(
        WorkbookErrorCode.FORMULA_CACHE_MISMATCH,
        f"{cell} does not contain a cached Excel date.",
    )


def preflight_xlsx_archive(content):
    """Inspect only the OOXML ZIP directory before openpyxl decompresses it."""

    if content.startswith(_ENCRYPTED_OLE_SIGNATURE):
        raise WorkbookContractError(
            WorkbookErrorCode.ENCRYPTED_XLSX,
            "Password-protected or encrypted workbooks are not supported.",
        )
    try:
        with ZipFile(BytesIO(content), "r", allowZip64=True) as archive:
            members = archive.infolist()
    except (BadZipFile, LargeZipFile, OSError, ValueError) as exc:
        raise WorkbookContractError(
            WorkbookErrorCode.INVALID_XLSX,
            "The uploaded file is not a valid XLSX workbook.",
        ) from exc

    if any(member.flag_bits & 0x1 for member in members):
        raise WorkbookContractError(
            WorkbookErrorCode.ENCRYPTED_XLSX,
            "Password-protected or encrypted ZIP members are not supported.",
        )
    member_count = len(members)
    total_uncompressed = sum(member.file_size for member in members)
    largest_member = max((member.file_size for member in members), default=0)
    if (
        member_count > MAX_ZIP_MEMBERS
        or total_uncompressed > MAX_ZIP_UNCOMPRESSED_BYTES
        or largest_member > MAX_ZIP_MEMBER_UNCOMPRESSED_BYTES
    ):
        raise WorkbookContractError(
            WorkbookErrorCode.RESOURCE_LIMIT,
            "The XLSX archive exceeds the supported resource limits.",
        )
    return XlsxArchiveResources(
        member_count=member_count,
        total_uncompressed_bytes=total_uncompressed,
        largest_member_bytes=largest_member,
    )


def _load_workbook_views(content):
    preflight_xlsx_archive(content)
    try:
        source = load_workbook(
            BytesIO(content), data_only=False, read_only=False, keep_links=False
        )
        cached = load_workbook(
            BytesIO(content), data_only=True, read_only=True, keep_links=False
        )
    except Exception as exc:
        raise WorkbookContractError(
            WorkbookErrorCode.INVALID_XLSX,
            "The uploaded file is not a readable XLSX workbook.",
        ) from exc
    return source, cached


def _validate_structure(source):
    if SUPPORTED_SHEET not in source.sheetnames:
        raise WorkbookContractError(
            WorkbookErrorCode.SHEET_MISSING,
            f"Required sheet {SUPPORTED_SHEET!r} is missing.",
        )
    if tuple(source.sheetnames) != EXPECTED_SHEET_NAMES:
        raise WorkbookContractError(
            WorkbookErrorCode.CONTRACT_MISMATCH,
            "Workbook sheet names or order do not match the supported contract.",
        )

    sheet = source[SUPPORTED_SHEET]
    if sheet["B2"].value != EXPECTED_TITLE:
        raise WorkbookContractError(
            WorkbookErrorCode.HEADER_MISMATCH,
            "Cell B2 does not match the supported 2026 master title.",
        )
    if (
        sheet["A3"].value != EXPECTED_DATE_HEADER
        or sheet["B3"].value != EXPECTED_WORSHIP_HEADER
    ):
        raise WorkbookContractError(
            WorkbookErrorCode.HEADER_MISMATCH,
            "Cells A3/B3 do not match the supported operational headers.",
        )
    merge_ranges = tuple(sorted(str(item) for item in sheet.merged_cells.ranges))
    if merge_ranges != EXPECTED_MERGES:
        raise WorkbookContractError(
            WorkbookErrorCode.CONTRACT_MISMATCH,
            "Merged cells do not match the supported workbook structure.",
        )

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and (cell.row > 58 or cell.column > 15):
                raise WorkbookContractError(
                    WorkbookErrorCode.CONTRACT_MISMATCH,
                    "Operational content exists outside the supported A1:O58 region.",
                )
    return sheet


def _validate_informational_rows(source_sheet, cached_sheet):
    if source_sheet["A5"].value != "1/9/26(Fri)":
        raise WorkbookContractError(
            WorkbookErrorCode.CONTRACT_MISMATCH,
            "The known Friday informational row at A5 changed.",
        )
    for row_number, expected in ((57, date(2027, 1, 3)), (58, date(2027, 1, 10))):
        source_cell = source_sheet.cell(row_number, 1)
        cached_cell = cached_sheet.cell(row_number, 1)
        expected_formula = f"=A{row_number - 1}+7"
        if source_cell.data_type != "f" or str(source_cell.value).strip() != expected_formula:
            raise WorkbookContractError(
                WorkbookErrorCode.CONTRACT_MISMATCH,
                f"A{row_number} does not match the known 2027 spillover formula.",
            )
        if _normalize_excel_date(cached_cell.value, cell=f"A{row_number}") != expected:
            raise WorkbookContractError(
                WorkbookErrorCode.CONTRACT_MISMATCH,
                f"A{row_number} does not match the known 2027 spillover date.",
            )


def _parse_supported_date(source_sheet, cached_sheet, row_number):
    expected = _expected_date_for_row(row_number)
    source_cell = source_sheet.cell(row_number, 1)
    cached_cell = cached_sheet.cell(row_number, 1)
    cell_name = f"A{row_number}"

    if row_number == 4:
        if source_cell.data_type == "f":
            raise WorkbookContractError(
                WorkbookErrorCode.DATE_MISMATCH,
                "A4 must be the literal date anchor, not a formula.",
            )
        value = _normalize_excel_date(source_cell.value, cell=cell_name)
        if value != expected:
            raise WorkbookContractError(
                WorkbookErrorCode.DATE_MISMATCH,
                "A4 must be the literal date 2026-01-04.",
            )
        return value, "literal"

    prior_row = 4 if row_number == 6 else row_number - 1
    expected_formula = f"=A{prior_row}+7"
    if (
        source_cell.data_type != "f"
        or str(source_cell.value).strip() != expected_formula
    ):
        raise WorkbookContractError(
            WorkbookErrorCode.FORMULA_CACHE_MISMATCH,
            f"{cell_name} does not contain the exact allowed weekly formula.",
        )
    value = _normalize_excel_date(cached_cell.value, cell=cell_name)
    if value != expected or value.year != SUPPORTED_YEAR or value.weekday() != 6:
        raise WorkbookContractError(
            WorkbookErrorCode.FORMULA_CACHE_MISMATCH,
            f"{cell_name} cached date does not match the deterministic weekly sequence.",
        )
    return value, "formula_cached"


def _parse_token(cell):
    if cell.data_type == "f" or not isinstance(cell.value, str):
        raise WorkbookContractError(
            WorkbookErrorCode.UNSUPPORTED_TOKEN,
            f"{cell.coordinate} must contain a literal supported rotation token.",
        )
    match = _TOKEN_RE.match(cell.value.strip())
    if match is None:
        raise WorkbookContractError(
            WorkbookErrorCode.UNSUPPORTED_TOKEN,
            f"{cell.coordinate} does not start with an allowed A/C1/C2/C3 token.",
        )
    return match.group(1)


def _actual_token_counts(rows):
    counts = Counter(row.token for row in rows)
    if set(counts) - set(TOKEN_ORDER):
        raise ValueError("Unsupported token in normalized workbook rows.")
    return {token: counts[token] for token in TOKEN_ORDER if counts[token]}


def parse_known_worship_workbook(content, *, filename="workbook.xlsx"):
    """Parse only the frozen 2026 SVCA Bethany 09:30 workbook contract."""

    if not isinstance(content, bytes) or not content:
        raise WorkbookContractError(
            WorkbookErrorCode.INVALID_XLSX, "The uploaded workbook is empty."
        )
    if len(content) > MAX_UPLOAD_BYTES:
        raise WorkbookContractError(
            WorkbookErrorCode.INVALID_XLSX, "The uploaded workbook is too large."
        )

    source, cached = _load_workbook_views(content)
    try:
        source_sheet = _validate_structure(source)
        cached_sheet = cached[SUPPORTED_SHEET]
        _validate_informational_rows(source_sheet, cached_sheet)
        rows = []
        token_counts = Counter()
        for row_number in SUPPORTED_ROWS:
            local_date, date_kind = _parse_supported_date(
                source_sheet, cached_sheet, row_number
            )
            token = _parse_token(source_sheet.cell(row_number, 2))
            token_counts[token] += 1
            rows.append(
                ParsedWorkbookRow(
                    source_row=row_number,
                    source_cell=f"A{row_number}/B{row_number}",
                    local_date=local_date,
                    date_kind=date_kind,
                    token=token,
                )
            )
        return ParsedWorshipWorkbook(
            filename=_safe_filename(filename),
            sha256=sha256(content).hexdigest().upper(),
            rows=tuple(rows),
            token_counts={
                token: token_counts[token]
                for token in TOKEN_ORDER
                if token_counts[token]
            },
            unsupported_rows=(
                {"row": 5, "kind": "friday", "date": "2026-01-09"},
                {"row": 57, "kind": "spillover", "date": "2027-01-03"},
                {"row": 58, "kind": "spillover", "date": "2027-01-10"},
            ),
        )
    finally:
        source.close()
        cached.close()


def _parsed_payload(parsed, *, user):
    return {
        "contract_revision": CONTRACT_REVISION,
        "signing_version": SIGNING_VERSION,
        "state_type": "parsed_workbook",
        "generated_at": timezone.now().isoformat(),
        "user_id": user.pk,
        "filename": parsed.filename,
        "sha256": parsed.sha256,
        "supported_sheet": SUPPORTED_SHEET,
        "rows": [
            {
                "source_row": row.source_row,
                "source_cell": row.source_cell,
                "local_date": row.local_date.isoformat(),
                "date_kind": row.date_kind,
                "token": row.token,
            }
            for row in parsed.rows
        ],
        "token_counts": parsed.token_counts,
    }


def sign_parsed_workbook(parsed, *, user):
    return signing.dumps(
        _parsed_payload(parsed, user=user), compress=True, salt=SIGNING_SALT
    )


def _decode_payload(token, *, user, state_type, max_age):
    try:
        payload = signing.loads(token, salt=SIGNING_SALT, max_age=max_age)
    except signing.BadSignature as exc:
        raise SignedWorkbookStateError("Signed preview state is invalid or expired.") from exc
    if (
        not isinstance(payload, dict)
        or payload.get("contract_revision") != CONTRACT_REVISION
        or payload.get("signing_version") != SIGNING_VERSION
        or payload.get("state_type") != state_type
        or payload.get("user_id") != getattr(user, "pk", None)
    ):
        raise SignedWorkbookStateError("Signed preview state does not match this user.")
    return payload


def decode_parsed_workbook(
    token, *, user, max_age=SIGNING_MAX_AGE_SECONDS
):
    payload = _decode_payload(
        token, user=user, state_type="parsed_workbook", max_age=max_age
    )
    rows_payload = payload.get("rows")
    if not isinstance(rows_payload, list) or len(rows_payload) != 52:
        raise SignedWorkbookStateError("Signed workbook rows are malformed.")
    rows = []
    try:
        for item in rows_payload:
            if set(item) != {
                "source_row",
                "source_cell",
                "local_date",
                "date_kind",
                "token",
            }:
                raise ValueError
            row = ParsedWorkbookRow(
                source_row=int(item["source_row"]),
                source_cell=str(item["source_cell"]),
                local_date=date.fromisoformat(item["local_date"]),
                date_kind=str(item["date_kind"]),
                token=str(item["token"]),
            )
            rows.append(row)
    except (TypeError, ValueError, KeyError) as exc:
        raise SignedWorkbookStateError("Signed workbook rows are malformed.") from exc
    try:
        actual_token_counts = _actual_token_counts(rows)
    except ValueError as exc:
        raise SignedWorkbookStateError(
            "Signed workbook contract facts are invalid."
        ) from exc
    payload_token_counts = payload.get("token_counts")
    token_counts_are_canonical = (
        isinstance(payload_token_counts, dict)
        and all(
            token in TOKEN_ORDER
            and type(count) is int
            and count > 0
            for token, count in payload_token_counts.items()
        )
        and payload_token_counts == actual_token_counts
        and sum(actual_token_counts.values()) == len(SUPPORTED_ROWS)
    )
    row_semantics_are_canonical = all(
        row.source_cell == f"A{row.source_row}/B{row.source_row}"
        and row.local_date == _expected_date_for_row(row.source_row)
        and (
            row.date_kind == "literal"
            if row.source_row == 4
            else row.date_kind == "formula_cached"
        )
        for row in rows
    )
    if (
        tuple(row.source_row for row in rows) != SUPPORTED_ROWS
        or not row_semantics_are_canonical
        or not token_counts_are_canonical
    ):
        raise SignedWorkbookStateError("Signed workbook contract facts are invalid.")
    return ParsedWorshipWorkbook(
        filename=_safe_filename(payload.get("filename")),
        sha256=str(payload.get("sha256")),
        rows=tuple(rows),
        token_counts=actual_token_counts,
        unsupported_rows=(
            {"row": 5, "kind": "friday", "date": "2026-01-09"},
            {"row": 57, "kind": "spillover", "date": "2027-01-03"},
            {"row": 58, "kind": "spillover", "date": "2027-01-10"},
        ),
    )


def _local_year_bounds():
    local_tz = timezone.get_current_timezone()
    return (
        timezone.make_aware(datetime(2026, 1, 1), local_tz),
        timezone.make_aware(datetime(2027, 1, 1), local_tz),
    )


def match_exact_service_event_targets(parsed):
    """Classify exact existing profile targets without creating or changing rows."""

    start, end = _local_year_bounds()
    events = list(
        ServiceEvent.objects.filter(start_datetime__gte=start, start_datetime__lt=end)
        .select_related("rotation_anchor_team", "host_language_unit")
        .prefetch_related(
            "audience_scope_links__unit",
            "required_team_links__ministry_team",
        )
        .order_by("start_datetime", "id")
    )
    events_by_date = defaultdict(list)
    for event in events:
        events_by_date[timezone.localtime(event.start_datetime).date()].append(event)

    matches = []
    for row in parsed.rows:
        same_date = events_by_date[row.local_date]
        exact = []
        for event in same_date:
            local_start = timezone.localtime(event.start_datetime)
            if (
                event.service_profile_key == SUPPORTED_PROFILE_KEY
                and event.event_type == SUPPORTED_EVENT_TYPE
                and local_start.time().replace(tzinfo=None) == SUPPORTED_LOCAL_TIME
            ):
                exact.append(event)
        parallel_count = len(same_date) - len(exact)
        if not exact:
            matches.append(
                TargetMatch(
                    row=row,
                    state=TargetMatchState.NO_TARGET,
                    event=None,
                    exact_target_ids=(),
                    audience_readiness=None,
                    parallel_evidence_count=parallel_count,
                )
            )
            continue
        if len(exact) > 1:
            matches.append(
                TargetMatch(
                    row=row,
                    state=TargetMatchState.MULTIPLE_EXACT_TARGETS,
                    event=None,
                    exact_target_ids=tuple(item.pk for item in exact),
                    audience_readiness=None,
                    parallel_evidence_count=parallel_count,
                )
            )
            continue
        event = exact[0]
        audience = service_event_audience_readiness(event)
        if event.status not in {
            ServiceEvent.STATUS_PUBLISHED,
            ServiceEvent.STATUS_COMPLETED,
        }:
            state = TargetMatchState.LIFECYCLE_CONFLICT
        elif not audience["ready"]:
            state = TargetMatchState.AUDIENCE_INVALID_CONFLICT
        else:
            state = TargetMatchState.EXACT_TARGET_MATCHED
        matches.append(
            TargetMatch(
                row=row,
                state=state,
                event=event,
                exact_target_ids=(event.pk,),
                audience_readiness=audience,
                parallel_evidence_count=parallel_count,
            )
        )
    return tuple(matches)


def mapping_candidate_teams(parsed, *, matches=None):
    """Return per-token unions of current canonical destination candidates."""

    matches = tuple(matches or match_exact_service_event_targets(parsed))
    candidates = {token: {} for token in parsed.token_counts}
    for match in matches:
        if match.state != TargetMatchState.EXACT_TARGET_MATCHED:
            continue
        inspection = inspect_worship_ownership_consistency(match.event)
        for candidate in inspection.eligible_candidates:
            candidates[match.row.token][candidate.team.pk] = candidate.team
    return {
        token: tuple(
            sorted(
                teams.values(),
                key=lambda team: (team.name, team.name_en, team.pk),
            )
        )
        for token, teams in candidates.items()
    }


def _target_blocker(target_state):
    return {
        TargetMatchState.NO_TARGET: PreviewBlocker.TARGET_MISSING,
        TargetMatchState.MULTIPLE_EXACT_TARGETS: PreviewBlocker.TARGET_AMBIGUOUS,
        TargetMatchState.LIFECYCLE_CONFLICT: PreviewBlocker.TARGET_LIFECYCLE,
        TargetMatchState.AUDIENCE_INVALID_CONFLICT: PreviewBlocker.TARGET_AUDIENCE,
    }.get(target_state)


def _validated_mapping(parsed, mapping):
    resolved = {}
    for token in parsed.token_counts:
        team = mapping.get(token)
        if team is not None and (not team.is_active or not team.is_assignable):
            raise MappingValidationError(
                f"Mapping for token {token} must use an active assignable team."
            )
        resolved[token] = team
    return resolved


def build_worship_import_preview(*, parsed, mapping, user):
    """Build a current, privacy-bounded, signed, zero-write preview."""

    mappings = _validated_mapping(parsed, mapping)
    matches = match_exact_service_event_targets(parsed)
    rows = []
    normalized_rows = []
    conflict_states = {
        WorshipOwnershipConsistencyState.INVALID_SELECTION,
        WorshipOwnershipConsistencyState.OFF_TEAM_CONFLICT,
        WorshipOwnershipConsistencyState.OUT_OF_SCOPE_WORSHIP_CONFLICT,
        WorshipOwnershipConsistencyState.MULTIPLE_CURRENT_WORSHIP_ASSIGNMENTS,
        WorshipOwnershipConsistencyState.DUPLICATE_SELECTED_TEAM_ASSIGNMENT,
    }

    for match in matches:
        proposed_team = mappings.get(match.row.token)
        event = match.event
        current_team = event.rotation_anchor_team if event is not None else None
        ownership_state = None
        downstream = ()
        fingerprints = None
        blocker = _target_blocker(match.state)
        classification = PreviewClassification.BLOCKED

        if match.state == TargetMatchState.EXACT_TARGET_MATCHED:
            inspection = inspect_worship_ownership_consistency(event)
            ownership_state = inspection.state
            eligible_team_ids = {
                candidate.team.pk for candidate in inspection.eligible_candidates
            }
            if proposed_team is None:
                blocker = PreviewBlocker.MAPPING_UNRESOLVED
            elif proposed_team.pk not in eligible_team_ids:
                blocker = PreviewBlocker.TEAM_INELIGIBLE
            elif inspection.state in conflict_states:
                blocker = PreviewBlocker.OWNERSHIP_CONFLICT
            elif (
                event.rotation_anchor_team_id != proposed_team.pk
                and inspection.current_worship_assignments
            ):
                blocker = PreviewBlocker.CURRENT_WORSHIP_ASSIGNMENT
            elif event.rotation_anchor_team_id == proposed_team.pk:
                classification = PreviewClassification.NO_OP
                blocker = None
            else:
                classification = PreviewClassification.PROPOSED_CHANGE
                blocker = None

            downstream, downstream_fingerprint = build_worship_downstream_projection(
                event
            )
            fingerprints = build_worship_preview_fingerprints(
                event, inspection, downstream_fingerprint
            )
            if classification != PreviewClassification.PROPOSED_CHANGE:
                downstream = ()

        row = WorshipImportPreviewRow(
            source=match.row,
            target_state=match.state,
            event=event,
            current_team=current_team,
            proposed_team=proposed_team,
            classification=classification,
            blocker=blocker,
            ownership_state=ownership_state,
            downstream_impacts=tuple(downstream),
            parallel_evidence_count=match.parallel_evidence_count,
            fingerprints=fingerprints,
        )
        rows.append(row)
        normalized_rows.append(
            {
                "source_row": match.row.source_row,
                "source_cell": match.row.source_cell,
                "local_date": match.row.local_date.isoformat(),
                "date_kind": match.row.date_kind,
                "token": match.row.token,
                "target_state": match.state.value,
                "event_id": getattr(event, "pk", None),
                "current_team_id": getattr(current_team, "pk", None),
                "proposed_team_id": getattr(proposed_team, "pk", None),
                "classification": classification.value,
                "blocker": blocker.value if blocker else None,
                "ownership_state": (
                    ownership_state.value if ownership_state is not None else None
                ),
                "parallel_evidence_count": match.parallel_evidence_count,
                "fingerprints": fingerprints,
            }
        )

    payload = {
        "contract_revision": CONTRACT_REVISION,
        "signing_version": SIGNING_VERSION,
        "state_type": "normalized_preview",
        "generated_at": timezone.now().isoformat(),
        "user_id": user.pk,
        "filename": parsed.filename,
        "sha256": parsed.sha256,
        "supported_sheet": SUPPORTED_SHEET,
        "mapping_team_ids": {
            token: getattr(mappings[token], "pk", None)
            for token in parsed.token_counts
        },
        "rows": normalized_rows,
    }
    signed_payload = signing.dumps(payload, compress=True, salt=SIGNING_SALT)
    return WorshipImportPreview(
        parsed=parsed,
        mappings=mappings,
        rows=tuple(rows),
        normalized_payload=payload,
        signed_payload=signed_payload,
    )


def decode_signed_worship_import_preview(
    token, *, user, max_age=SIGNING_MAX_AGE_SECONDS
):
    payload = _decode_payload(
        token, user=user, state_type="normalized_preview", max_age=max_age
    )
    rows = payload.get("rows")
    mappings = payload.get("mapping_team_ids")
    if (
        not isinstance(rows, list)
        or len(rows) != 52
        or not isinstance(mappings, dict)
    ):
        raise SignedWorkbookStateError("Signed normalized preview is malformed.")
    allowed_row_keys = {
        "source_row",
        "source_cell",
        "local_date",
        "date_kind",
        "token",
        "target_state",
        "event_id",
        "current_team_id",
        "proposed_team_id",
        "classification",
        "blocker",
        "ownership_state",
        "parallel_evidence_count",
        "fingerprints",
    }
    if any(not isinstance(row, dict) or set(row) != allowed_row_keys for row in rows):
        raise SignedWorkbookStateError("Signed normalized preview is malformed.")
    if any(not isinstance(row["token"], str) for row in rows):
        raise SignedWorkbookStateError("Signed normalized preview is malformed.")
    normalized_counts = Counter(row["token"] for row in rows)
    if (
        set(normalized_counts) - set(TOKEN_ORDER)
        or sum(normalized_counts.values()) != len(SUPPORTED_ROWS)
        or set(mappings)
        != {token for token in TOKEN_ORDER if normalized_counts[token]}
        or any(
            value is not None and type(value) is not int
            for value in mappings.values()
        )
    ):
        raise SignedWorkbookStateError("Signed normalized preview is malformed.")
    return payload
